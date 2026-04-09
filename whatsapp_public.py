import time
import random
import re
import json
import os
import ctypes
from datetime import datetime
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.action_chains import ActionChains
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ══════════════════════════════════════════════════════
#   SETTINGS
# ══════════════════════════════════════════════════════
WEBSITE_URL   = "yourwebsite.com"          # ← put your website here
MAX_CONTACTS  = 1000
CONTACTS_FILE = r"C:\Users\YOUR_USERNAME\whatsapp_contacts.xlsx"  # ← replace YOUR_USERNAME
PROGRESS_FILE = r"C:\Users\YOUR_USERNAME\whatsapp_progress.json"  # ← replace YOUR_USERNAME

BURST_PATTERN = [
    (3,  60,   180),
    (3,  180,  420),
    (3,  420,  720),
    (3,  720,  1200),
    (3,  1200, 1800),
]

# ══════════════════════════════════════════════════════
#   MESSAGE VARIANTS
# ══════════════════════════════════════════════════════
VARIANTS_WITH_NAME = [
    "Hi {name}! 👋\n\nDon't let a busy schedule cost you your degree or certification. 🎓\nWe quietly handle it for you — 100% discreet.\n\nReply 1 → Certification\nReply 2 → Exam\nReply 3 → Online Class\n\nDM anytime or visit: {url} 🙌",
    "Hey {name}! 😊\n\nHectic routine getting in the way of your studies?\nWe handle certifications, exams & online classes — professionally & discreetly. ✅\n\nReply 1 → Certification\nReply 2 → Exam\nReply 3 → Online Class\n\nLearn more: {url}",
    "Hi {name}! 📚\n\nBehind on exams or certifications because of work & life?\nLet us quietly take care of it — 100% confidential.\n\nReply 1 → Certification\nReply 2 → Exam\nReply 3 → Online Class\n\nDM or visit: {url} 🙌",
]

VARIANTS_WITHOUT_NAME = [
    "Hey! 👋\n\nDon't let a busy schedule cost you your degree or certification. 🎓\nWe quietly handle it for you — 100% discreet.\n\nReply 1 → Certification\nReply 2 → Exam\nReply 3 → Online Class\n\nDM anytime or visit: {url} 🙌",
    "Hi there! 😊\n\nHectic routine getting in the way of your studies?\nWe handle certifications, exams & online classes — professionally & discreetly. ✅\n\nReply 1 → Certification\nReply 2 → Exam\nReply 3 → Online Class\n\nLearn more: {url}",
    "Hello! 📚\n\nBehind on exams or certifications because of work & life?\nLet us quietly take care of it — 100% confidential.\n\nReply 1 → Certification\nReply 2 → Exam\nReply 3 → Online Class\n\nDM or visit: {url} 🙌",
]

# ══════════════════════════════════════════════════════
#   FAKE NAME DETECTION
# ══════════════════════════════════════════════════════
FAKE_NAME_WORDS = {
    "supplier", "client", "customer", "guy", "old", "new", "my", "no",
    "group", "class", "team", "shop", "store", "office", "contact",
    "number", "phone", "temp", "unknown", "test", "dubai", "karachi",
    "lahore", "islamabad", "bhai", "uncle", "aunty", "friend", "work",
    "business", "agent", "dealer", "buyer", "seller", "ref", "help",
    "support", "personal", "home", "main", "backup", "second", "other",
    "us", "uk", "uae", "usa", "canada", "tello", "jazz", "zong",
    "ufone", "telenor", "accountant", "driver", "manager", "acct",
    "busa", "psi", "exam", "biology", "azure", "salesforce", "proctor",
    "esim", "sim", "ref", "re", "wgu"
}

def classify_contact(name):
    cleaned = re.sub(r'[\s\-\(\)\+]', '', name)
    if cleaned.isdigit() and len(cleaned) >= 7:
        return "Phone No."
    if any(char.isdigit() for char in name):
        return "Label/Group"
    if len(name.strip().split()) > 3:
        return "Label/Group"
    first_word = name.strip().split()[0]
    if not first_word[0].isupper():
        return "Label/Group"
    if not first_word.isalpha():
        return "Label/Group"
    if len(first_word) < 2 or len(first_word) > 12:
        return "Label/Group"
    if first_word.lower() in FAKE_NAME_WORDS:
        return "Label/Group"
    return "Real Name"

# ══════════════════════════════════════════════════════
#   SLEEP PREVENTION
# ══════════════════════════════════════════════════════
def prevent_sleep():
    ctypes.windll.kernel32.SetThreadExecutionState(0x80000002)

def allow_sleep():
    ctypes.windll.kernel32.SetThreadExecutionState(0x80000000)

# ══════════════════════════════════════════════════════
#   PROGRESS
# ══════════════════════════════════════════════════════
def load_progress():
    if os.path.exists(PROGRESS_FILE):
        with open(PROGRESS_FILE, 'r') as f:
            return json.load(f)
    return {"sent": [], "failed": [], "last_run": None,
            "campaign_started": None, "campaign_finished": None,
            "total_sessions": 0}

def save_progress(progress):
    with open(PROGRESS_FILE, 'w') as f:
        json.dump(progress, f, indent=2)

def reset_progress():
    fresh = {"sent": [], "failed": [], "last_run": None,
             "campaign_started": datetime.now().strftime("%Y-%m-%d %H:%M"),
             "campaign_finished": None, "total_sessions": 0}
    save_progress(fresh)
    return fresh

# ══════════════════════════════════════════════════════
#   EXCEL — SAVE CONTACTS
# ══════════════════════════════════════════════════════
def save_contacts_to_excel(contacts):
    wb = Workbook()
    ws = wb.active
    ws.title = "Contacts"

    # Styles
    header_font    = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    header_fill    = PatternFill("solid", start_color="1F4E79")
    center_align   = Alignment(horizontal="center", vertical="center")
    border_side    = Side(style="thin", color="CCCCCC")
    cell_border    = Border(left=border_side, right=border_side,
                            top=border_side, bottom=border_side)

    green_fill  = PatternFill("solid", start_color="E2EFDA")
    orange_fill = PatternFill("solid", start_color="FCE4D6")
    blue_fill   = PatternFill("solid", start_color="DEEAF1")

    # Headers
    headers = ["#", "Name / Label", "Type", "Send? (delete row to skip)"]
    col_widths = [5, 35, 15, 30]

    for col, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = cell_border
        ws.column_dimensions[cell.column_letter].width = width

    ws.row_dimensions[1].height = 25

    # Data rows
    for i, name in enumerate(contacts, 1):
        contact_type = classify_contact(name)

        row = [i, name, contact_type, "✅ Will Send"]
        for col, value in enumerate(row, 1):
            cell = ws.cell(row=i + 1, column=col, value=value)
            cell.font = Font(name="Arial", size=10)
            cell.alignment = Alignment(vertical="center",
                                       horizontal="center" if col != 2 else "left")
            cell.border = cell_border

            # Color code by type
            if contact_type == "Real Name":
                cell.fill = green_fill
            elif contact_type == "Phone No.":
                cell.fill = blue_fill
            else:
                cell.fill = orange_fill

    # Freeze header row
    ws.freeze_panes = "A2"

    # Summary at bottom
    ws.cell(row=len(contacts) + 3, column=1,
            value=f"Total: {len(contacts)} contacts").font = Font(bold=True)
    ws.cell(row=len(contacts) + 4, column=1,
            value="⚠️  DELETE any rows you don't want to message, then save.").font = \
        Font(bold=True, color="FF0000")

    wb.save(CONTACTS_FILE)
    print(f"\n   ✅ Saved to: {CONTACTS_FILE}")
    print(f"   📊 Total contacts: {len(contacts)}")
    print(f"\n   🟢 Green  = Real Name")
    print(f"   🔵 Blue   = Phone Number")
    print(f"   🟠 Orange = Label/Group")
    print(f"\n   ➡️  Open the Excel file, DELETE rows you don't want,")
    print(f"      save the file, then run this script again → choose Option 2")

# ══════════════════════════════════════════════════════
#   EXCEL — LOAD CONTACTS
# ══════════════════════════════════════════════════════
def load_contacts_from_excel():
    if not os.path.exists(CONTACTS_FILE):
        print(f"\n   ❌ Contacts file not found: {CONTACTS_FILE}")
        print(f"   Run Option 1 first to scrape contacts!")
        return []

    wb = load_workbook(CONTACTS_FILE)
    ws = wb.active
    contacts = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[1] and str(row[1]).strip():
            name = str(row[1]).strip()
            # Skip summary rows
            if name.startswith("Total:") or name.startswith("⚠️"):
                continue
            contacts.append(name)

    print(f"   ✅ Loaded {len(contacts)} contacts from Excel")
    return contacts

# ══════════════════════════════════════════════════════
#   CHROME DRIVER
# ══════════════════════════════════════════════════════
def start_browser():
    options = webdriver.ChromeOptions()
    options.add_argument(r"--user-data-dir=C:\Users\YOUR_USERNAME\whatsapp_session")  # ← replace YOUR_USERNAME
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")
    options.add_argument("--disable-gpu")
    driver = webdriver.Chrome(options=options)
    driver.get("https://web.whatsapp.com")
    return driver

# ══════════════════════════════════════════════════════
#   SCRAPE CONTACTS
# ══════════════════════════════════════════════════════
def scrape_contacts(driver):
    print("\n   📋 Scraping contacts — scrolling to load all chats...")
    wait = WebDriverWait(driver, 20)
    wait.until(EC.presence_of_element_located(
        (By.XPATH, "//div[@aria-label='Chat list']")
    ))
    time.sleep(3)

    contacts = []
    seen = set()
    chat_list = driver.find_element(By.XPATH, "//div[@aria-label='Chat list']")
    last_count = 0
    no_change = 0

    while True:
        items = driver.find_elements(
            By.XPATH,
            "//div[@aria-label='Chat list']//span[@dir='auto' and @title]"
        )
        for item in items:
            title = item.get_attribute("title").strip()
            if title and title not in seen:
                seen.add(title)
                contacts.append(title)

        driver.execute_script("arguments[0].scrollTop += 800", chat_list)
        time.sleep(1.5)

        if len(contacts) == last_count:
            no_change += 1
            if no_change >= 5:
                break
        else:
            no_change = 0
        last_count = len(contacts)

        if len(contacts) >= MAX_CONTACTS:
            break

        print(f"   📋 Found {len(contacts)} contacts so far...", end="\r")

    print(f"   ✅ Total scraped: {len(contacts)} contacts          ")
    return contacts

# ══════════════════════════════════════════════════════
#   BUILD MESSAGE
# ══════════════════════════════════════════════════════
def build_message(contact):
    variant = random.randint(0, 2)
    contact_type = classify_contact(contact)
    if contact_type == "Real Name":
        first_name = contact.strip().split()[0]
        return VARIANTS_WITH_NAME[variant].format(name=first_name, url=WEBSITE_URL), variant + 1
    return VARIANTS_WITHOUT_NAME[variant].format(url=WEBSITE_URL), variant + 1

# ══════════════════════════════════════════════════════
#   SEND MESSAGE
# ══════════════════════════════════════════════════════
def send_to_contact(driver, contact, message):
    wait = WebDriverWait(driver, 20)
    driver.get("https://web.whatsapp.com")
    time.sleep(3)

    cleaned = re.sub(r'[\s\-\(\)\+]', '', contact)
    if cleaned.isdigit() and len(cleaned) >= 7:
        driver.get(f"https://web.whatsapp.com/send?phone={contact}")
        time.sleep(4)
    else:
        try:
            chat_item = wait.until(EC.element_to_be_clickable(
                (By.XPATH, f"//span[@title='{contact}']")
            ))
            chat_item.click()
            time.sleep(2)
        except:
            search = wait.until(EC.element_to_be_clickable(
                (By.XPATH, "//div[@contenteditable='true'][@data-tab='3']")
            ))
            search.click()
            time.sleep(1)
            search.send_keys(contact)
            time.sleep(3)
            result = wait.until(EC.element_to_be_clickable(
                (By.XPATH, f"//span[@title='{contact}']")
            ))
            result.click()
            time.sleep(2)

    msg_box = wait.until(EC.element_to_be_clickable(
        (By.XPATH, "//div[@contenteditable='true'][@data-tab='10']")
    ))
    msg_box.click()
    time.sleep(1)

    lines = message.split('\n')
    for i, line in enumerate(lines):
        msg_box.send_keys(line)
        if i < len(lines) - 1:
            ActionChains(driver).key_down('\ue008').send_keys('\ue007').key_up('\ue008').perform()
            time.sleep(0.1)

    time.sleep(1)
    send_btn = wait.until(EC.element_to_be_clickable(
        (By.XPATH, "//button[@aria-label='Send']")
    ))
    send_btn.click()
    time.sleep(3)

# ══════════════════════════════════════════════════════
#   OPTION 1 — SCRAPE TO EXCEL
# ══════════════════════════════════════════════════════
def option_scrape():
    print("\n" + "=" * 55)
    print("   STEP 1 — Scraping contacts to Excel")
    print("=" * 55)

    driver = start_browser()
    print("\n   Scan QR code if needed, then press Enter...")
    input()

    contacts = scrape_contacts(driver)
    driver.quit()

    save_contacts_to_excel(contacts)

# ══════════════════════════════════════════════════════
#   OPTION 2 — SEND FROM EXCEL
# ══════════════════════════════════════════════════════
def option_send():
    print("\n" + "=" * 55)
    print("   STEP 2 — Sending messages from Excel")
    print("=" * 55)

    all_contacts = load_contacts_from_excel()
    if not all_contacts:
        return

    progress = load_progress()

    # Check if campaign complete
    already_sent = set(progress["sent"])
    pending = [c for c in all_contacts if c not in already_sent]

    if not pending:
        print("\n   ✅ All contacts in the Excel have been messaged!")
        print("   Started  :", progress.get("campaign_started", "N/A"))
        print("   Finished :", progress.get("campaign_finished", "N/A"))
        print("\n   Do you want to start a FRESH campaign? (e.g. next month)")
        while True:
            ans = input("   Type YES to reset, or NO to exit: ").strip().lower()
            if ans == "yes":
                progress = reset_progress()
                pending  = all_contacts
                print("   🗑️  Progress reset — starting fresh!")
                break
            elif ans == "no":
                print("   👋 Come back next month!")
                return
            else:
                print("   Please type YES or NO")
    else:
        if progress["campaign_started"]:
            print(f"\n   ▶️  Resuming campaign")
            print(f"   Already sent : {len(already_sent)}")
            print(f"   Remaining    : {len(pending)}")
        else:
            progress["campaign_started"] = datetime.now().strftime("%Y-%m-%d %H:%M")
            save_progress(progress)
            print(f"\n   🆕 Starting new campaign — {len(pending)} contacts")

    # Preview
    print(f"\n{'#':<4} {'Contact':<30} {'Type':<15}")
    print("-" * 52)
    for i, c in enumerate(pending[:15]):
        ctype = classify_contact(c)
        icon  = "✅" if ctype == "Real Name" else ("📱" if ctype == "Phone No." else "⚠️ ")
        print(f"{i+1:<4} {c:<30} {icon} {ctype}")
    if len(pending) > 15:
        print(f"     ... and {len(pending) - 15} more")
    print("-" * 52)

    print(f"\n⚠️  Ready to send to {len(pending)} contacts.")
    print("Press Enter to START, or Ctrl+C to CANCEL...")
    input()

    prevent_sleep()
    driver = start_browser()
    print("\n   Scan QR if needed, then press Enter...")
    input()

    progress["total_sessions"] += 1
    sent_count  = 0
    fail_count  = 0
    burst_index = 0
    burst_count = 0
    burst_size, burst_min, burst_max = BURST_PATTERN[burst_index]

    print(f"\n🚀 Starting — {len(pending)} contacts to go!\n")

    for i, contact in enumerate(pending):
        message, variant_num = build_message(contact)
        ctype = classify_contact(contact)
        icon  = "✅" if ctype == "Real Name" else ("📱" if ctype == "Phone No." else "⚠️ ")

        print(f"[{i+1}/{len(pending)}] {icon} {contact}  (Variant {variant_num})")

        try:
            send_to_contact(driver, contact, message)
            progress["sent"].append(contact)
            progress["last_run"] = datetime.now().strftime("%Y-%m-%d %H:%M")
            save_progress(progress)
            sent_count += 1
            print(f"   ✅ Sent!  |  Today: {sent_count}  |  Left: {len(pending)-(i+1)}")
        except Exception as e:
            progress["failed"].append(contact)
            save_progress(progress)
            fail_count += 1
            print(f"   ❌ Failed — skipping")

        if i >= len(pending) - 1:
            break

        burst_count += 1
        if burst_count >= burst_size:
            burst_count = 0
            burst_index = (burst_index + 1) % len(BURST_PATTERN)
            burst_size, burst_min, burst_max = BURST_PATTERN[burst_index]
            print(f"\n   🔄 Burst level {burst_index + 1}")

        delay = random.randint(burst_min, burst_max)
        print(f"   ⏱️  Waiting {delay//60}m {delay%60}s...\n")
        time.sleep(delay)

    if len(progress["sent"]) >= len(all_contacts):
        progress["campaign_finished"] = datetime.now().strftime("%Y-%m-%d %H:%M")
        save_progress(progress)

    print("\n" + "=" * 55)
    print("   SESSION COMPLETE")
    print("=" * 55)
    print(f"   ✅ Sent this session : {sent_count}")
    print(f"   ❌ Failed            : {fail_count}")
    print(f"   📊 Total ever sent   : {len(progress['sent'])}")
    remaining = len(all_contacts) - len(progress['sent'])
    if remaining > 0:
        print(f"   📋 Still remaining  : {remaining}")
        print(f"   ▶️  Run again to continue!")
    else:
        print(f"   🏁 Campaign complete! Come back next month.")
    print("=" * 55)

    allow_sleep()
    driver.quit()

# ══════════════════════════════════════════════════════
#   MAIN MENU
# ══════════════════════════════════════════════════════
def main():
    while True:
        print("\n" + "=" * 55)
        print("   WhatsApp Auto Messenger")
        print("=" * 55)
        print("   1 → Scrape contacts to Excel")
        print("   2 → Send messages from Excel")
        print("   3 → Exit")
        print("=" * 55)
        choice = input("   Choose option (1/2/3): ").strip()

        if choice == "1":
            option_scrape()
        elif choice == "2":
            option_send()
        elif choice == "3":
            print("\n   👋 Goodbye!\n")
            break
        else:
            print("   Please enter 1, 2, or 3")

main()
